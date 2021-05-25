from openpyxl import Workbook
from bs4 import BeautifulSoup as bs

# 엑셀파일 쓰기
wb = Workbook()
ws = wb.active
ws.title = "제주편"

title = ['name','explain','division','call','locate','break','time_st','time_fi','menu','etc']


file_name = "D:/2021/greentable/data_modify.html"
page = open(file_name, 'r', encoding='utf-8').read()
soup = bs(page, 'html.parser')
dataset = [] #list

column = 1

for t in range(len(title)):
    write_ws = ws.cell(row=1, column=column)
    write_ws.value = title[t]
    column += 1

column = 1
for t in title:
    article = soup.find_all("span",{"class":t})
    for a in article:
        dataset.append(a.text)
    for i in range(len(dataset)):
        write_ws = ws.cell(row=i+2,column= column)
        write_ws.value = dataset[i]
    print(dataset)
    dataset.clear()
    column += 1

wb.save("D:/2021/greentable/data_veganism.xlsx")